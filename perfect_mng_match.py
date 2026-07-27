import openpyxl
import re

# Load Muro courses
muro_courses = {}
with open('muro_mng_courses.txt', 'r', encoding='utf-8') as f:
    for line in f:
        line = line.strip()
        if '|' in line:
            cid, title = line.split('|', 1)
            muro_courses[title.strip()] = cid.strip()

def normalize(text):
    text = text.lower().strip()
    text = text.replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return ' '.join(text.split())

# Create course mapping helper
muro_mapped = {}
# Excel courses unique names list
excel_unique_courses = set()

wb = openpyxl.load_workbook('dereceuzem s7.xlsx')
sheet = wb.active

for row in list(sheet.iter_rows(values_only=True))[1:]:
    if len(row) > 4 and row[1] == 'Dereceuzem' and row[4]:
        excel_unique_courses.add(row[4].strip())

for ec in excel_unique_courses:
    norm_ec = normalize(ec)
    for mc, mc_id in muro_courses.items():
        if normalize(mc) == norm_ec or norm_ec in normalize(mc) or normalize(mc) in norm_ec:
            muro_mapped[ec] = mc_id
            break

print(f"Mapped {len(muro_mapped)} courses successfully.")

sql_statements = []
skipped_rows = 0
inserted_rows = 0

for idx, row in enumerate(list(sheet.iter_rows(values_only=True))[1:], 1):
    if len(row) < 10 or row[1] != 'Dereceuzem':
        skipped_rows += 1
        continue
    
    bbb_room_name = row[3].strip() if row[3] else "Canlı Ders Kaydı"
    excel_course_name = row[4].strip() if row[4] else ""
    date_val = row[5] # datetime object or string
    folder_id = row[8].strip() if row[8] else ""
    
    if not folder_id or not excel_course_name:
        skipped_rows += 1
        continue
        
    course_id = muro_mapped.get(excel_course_name)
    if not course_id:
        print(f"Warning: Could not map course '{excel_course_name}' for row {idx}")
        skipped_rows += 1
        continue
        
    # Format date for postgres
    if hasattr(date_val, 'strftime'):
        date_str = date_val.strftime('%Y-%m-%d %H:%M:%S')
    else:
        date_str = str(date_val)
        
    # Escape single quotes in room title
    clean_room_title = bbb_room_name.replace("'", "''")
    
    # Construct VideoUrl using clean BBB 2.3+ format
    video_url = f"https://canli.mng.muro.click/playback/presentation/2.3/{folder_id}"
    
    sql = f"""DO $$ 
DECLARE
    sid uuid;
BEGIN
    sid := gen_random_uuid();
    IF NOT EXISTS (SELECT 1 FROM "Sessions" WHERE "CourseId" = '{course_id}' AND "BbbMeetingId" = '{folder_id}') THEN
        INSERT INTO "Sessions" ("Id", "CourseId", "Title", "BbbMeetingId", "VideoUrl", "ScheduledStart", "IsDeleted", "CreatedAt", "Status", "Order", "RecordingEnabled", "IsFree")
        VALUES (sid, '{course_id}', '{clean_room_title}', '{folder_id}', '{video_url}', '{date_str}', false, NOW(), 3, 0, true, false);
        
        INSERT INTO "CourseMedias" ("Id", "CourseId", "SessionId", "OrderIndex", "CreatedAt")
        VALUES (gen_random_uuid(), '{course_id}', sid, -1, NOW());
    END IF;
END $$;"""
    sql_statements.append(sql)
    inserted_rows += 1

output_sql_file = 'mng_sessions_insert.sql'
with open(output_sql_file, 'w', encoding='utf-8') as f:
    f.write("-- ==================================================\n")
    f.write("-- DERECE UZEM (mng) - SESSIONS & COURSE MEDIAS IMPORT\n")
    f.write(f"-- Total Sessions: {len(sql_statements)}\n")
    f.write("-- ==================================================\n\n")
    f.write('\n\n'.join(sql_statements))

print(f"Generated {inserted_rows} SQL insert statements in '{output_sql_file}'.")
print(f"Skipped rows: {skipped_rows}")
