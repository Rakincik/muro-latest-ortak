import os
import sys
import uuid
import datetime
import re
import argparse
import psycopg2
import openpyxl

def normalize(s):
    if not s:
        return ''
    s = str(s).lower().strip()
    s = s.replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
    return re.sub(r'[^a-z0-9]', '', s)

def clean_phone_number(phone):
    if not phone:
        return None
    phone_str = str(phone).strip()
    main_part = phone_str.split('.')[0].split(',')[0]
    digits = "".join(c for c in main_part if c.isdigit())
    while digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 12 and digits.startswith("90"):
        digits = digits[2:]
    return digits if digits else None

def parse_name(full_name):
    if not full_name:
        return "", ""
    parts = str(full_name).strip().split()
    if len(parts) == 0:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]

def main():
    parser = argparse.ArgumentParser(description="OMURHOCA LMS Importer")
    parser.add_argument("--host", help="PostgreSQL host")
    parser.add_argument("--port", type=int, help="PostgreSQL port")
    parser.add_argument("--dbname", help="PostgreSQL database name")
    parser.add_argument("--user", help="PostgreSQL username")
    parser.add_argument("--password", help="PostgreSQL password")
    parser.add_argument("--execute", action="store_true", help="Execute the database writes (defaults to dry-run)")
    args = parser.parse_args()

    print("=" * 60)
    print("        OMURHOCA LMS - COURSES, GROUPS & STUDENTS IMPORT")
    print("=" * 60)

    # Database parameters discovery (fallback to default)
    conn_params = {}
    
    # Try discovering appsettings.json connection string
    paths = [
        "src/MURO.API/appsettings.json",
        "appsettings.json",
        "../src/MURO.API/appsettings.json"
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    content = f.read()
                    match = re.search(r'"DefaultConnection"\s*:\s*"([^"]+)"', content)
                    if match:
                        conn_str = match.group(1)
                        for item in conn_str.split(';'):
                            if '=' in item:
                                k, v = item.split('=', 1)
                                k_lower = k.strip().lower()
                                v_val = v.strip()
                                if k_lower == 'host': conn_params['host'] = v_val
                                elif k_lower == 'port': conn_params['port'] = int(v_val)
                                elif k_lower == 'database': conn_params['dbname'] = v_val
                                elif k_lower == 'username': conn_params['user'] = v_val
                                elif k_lower == 'password': conn_params['password'] = v_val
            except Exception:
                pass

    # Override or set from args
    if args.host: conn_params['host'] = args.host
    if args.port: conn_params['port'] = args.port
    if args.dbname: conn_params['dbname'] = args.dbname
    if args.user: conn_params['user'] = args.user
    if args.password: conn_params['password'] = args.password

    # Default fallback
    if 'host' not in conn_params: conn_params['host'] = 'localhost'
    if 'port' not in conn_params: conn_params['port'] = 5432
    if 'dbname' not in conn_params: conn_params['dbname'] = 'muro_dev'
    if 'user' not in conn_params: conn_params['user'] = 'muro_user'
    if 'password' not in conn_params: conn_params['password'] = 'muro_pass_2024'

    # Resolve Excel file paths dynamically relative to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file1 = os.path.join(script_dir, "ömürhoca.xlsx")
    file2 = os.path.join(script_dir, "omurhoca_ogrenciler_hepsi.xlsx")

    # 1. READ OMERHOCA.XLSX
    print(f"[*] Reading '{file1}'...")
    if not os.path.exists(file1):
        print(f"[!] Error: File '{file1}' not found.")
        sys.exit(1)

    wb1 = openpyxl.load_workbook(file1, data_only=True)
    
    # Parse Sayfa1 (User lookup)
    s1 = wb1['Sayfa1']
    rows1 = list(s1.iter_rows(values_only=True))
    user_lookup = {}
    print(f"    - Sayfa1 rows: {len(rows1)}")
    # Headers at Row 2: (None, 'rol', 'adi_soyadi', 'telefon', 'eposta', 'gruplar')
    for r in rows1[3:]:
        rol = r[1]
        name = r[2]
        phone = r[3]
        email = r[4]
        if name:
            name_str = str(name).strip()
            cleaned_phone = clean_phone_number(phone)
            cleaned_email = str(email).strip().lower() if email else None
            if cleaned_email and "@" not in cleaned_email:
                cleaned_email = None
            user_lookup[normalize(name_str)] = {
                'phone': cleaned_phone,
                'email': cleaned_email,
                'role': str(rol).strip() if rol else 'Öğrenci'
            }
    print(f"    - User lookup build: {len(user_lookup)} users mapped from Sayfa1.")

    # Parse Sayfa2 Left & Right
    s2 = wb1['Sayfa2']
    rows2 = list(s2.iter_rows(values_only=True))
    print(f"    - Sayfa2 rows: {len(rows2)}")

    courses_to_create = set()
    groups_to_create = set()
    course_groups_to_create = [] # (group_name, course_name, is_online)
    all_students = {} # norm_name -> { name, phone, email, groups: set(), courses: set() }

    # Sayfa2 Left Table parses courses, groups, and mappings
    # Headers at Row 1: grup_id(1), grup_adi(2), ders(3), aciklama(4), mod(5)
    for r in rows2[2:]:
        g_id = r[1]
        g_name = r[2]
        c_name = r[3]
        mode = r[5]
        if g_name:
            g_name_str = str(g_name).strip()
            groups_to_create.add(g_name_str)
            if c_name and str(c_name).strip() != "Gösterilecek veri yok.":
                c_name_str = str(c_name).strip()
                courses_to_create.add(c_name_str)
                is_online = (str(mode).strip().lower() == "canlı")
                course_groups_to_create.append((g_name_str, c_name_str, is_online))

    # Sayfa2 Right Table parses student group memberships
    # Headers at Row 1: grup_id(9), grup_adi(10), ad_soyad(11), rol(12)
    for r in rows2[2:]:
        g_name = r[10]
        ad_soyad = r[11]
        if g_name and ad_soyad and "gösterilecek" not in str(ad_soyad).lower():
            g_name_str = str(g_name).strip()
            student_name_str = str(ad_soyad).strip()
            norm_name = normalize(student_name_str)

            if norm_name not in all_students:
                s1_info = user_lookup.get(norm_name, {})
                all_students[norm_name] = {
                    'name': student_name_str,
                    'phone': s1_info.get('phone', None),
                    'email': s1_info.get('email', None),
                    'groups': set(),
                    'courses': set()
                }
            all_students[norm_name]['groups'].add(g_name_str)

    # 2. READ OMURHOCA_OGRENCILER_HEPSI.XLSX
    print(f"[*] Reading '{file2}'...")
    if not os.path.exists(file2):
        print(f"[!] Error: File '{file2}' not found.")
        sys.exit(1)

    wb2 = openpyxl.load_workbook(file2, data_only=True)
    s_list = wb2['Ogrenci Listesi']
    rows_list = list(s_list.iter_rows(values_only=True))
    print(f"    - Ogrenci Listesi rows: {len(rows_list)}")

    # Headers at Row 0: Ad Soyad, TELEFON, MAIL, GRUP, DERS
    for r in rows_list[1:]:
        name = r[0]
        phone = r[1]
        email = r[2]
        grup_str = r[3]
        ders_str = r[4]

        if not name:
            continue

        name_str = str(name).strip()
        norm_name = normalize(name_str)

        cleaned_phone = clean_phone_number(phone)
        cleaned_email = str(email).strip().lower() if email else None
        if cleaned_email and "@" not in cleaned_email:
            cleaned_email = None

        if norm_name not in all_students:
            all_students[norm_name] = {
                'name': name_str,
                'phone': cleaned_phone,
                'email': cleaned_email,
                'groups': set(),
                'courses': set()
            }
        else:
            if cleaned_phone:
                all_students[norm_name]['phone'] = cleaned_phone
            if cleaned_email:
                all_students[norm_name]['email'] = cleaned_email

        # Add groups
        if grup_str:
            for g in str(grup_str).split(','):
                g_clean = g.strip()
                if g_clean:
                    groups_to_create.add(g_clean)
                    all_students[norm_name]['groups'].add(g_clean)

        # Add individual courses
        if ders_str:
            for c in str(ders_str).split(','):
                c_clean = c.strip()
                if c_clean:
                    courses_to_create.add(c_clean)
                    all_students[norm_name]['courses'].add(c_clean)

    print(f"\n[+] Parsed Excel Data successfully:")
    print(f"    - Unique Groups to create  : {len(groups_to_create)}")
    print(f"    - Unique Courses to create : {len(courses_to_create)}")
    print(f"    - Unique Students in total : {len(all_students)}")
    print(f"    - Course-Group mappings    : {len(course_groups_to_create)}")

    # Connect to Database
    print(f"\n[*] Connecting to PostgreSQL database...")
    try:
        conn = psycopg2.connect(**conn_params)
        cur = conn.cursor()
        print("[+] Connected successfully!")
    except Exception as e:
        print(f"[!] Database Connection Failed: {e}")
        sys.exit(1)

    try:
        # Load existing database records for duplicate prevention
        print("[*] Retrieving existing database records...")
        
        cur.execute('SELECT "Id", "Title" FROM "Courses" WHERE "IsDeleted" = False;')
        db_courses = {r[1].lower().strip(): r[0] for r in cur.fetchall()}
        
        cur.execute('SELECT "Id", "Name" FROM "Groups" WHERE "IsDeleted" = False;')
        db_groups = {r[1].lower().strip(): r[0] for r in cur.fetchall()}
        
        cur.execute('SELECT "Id", "Email", "Phone", "Username" FROM "Users";')
        db_users = cur.fetchall()
        db_users_by_email = {r[1].lower().strip(): r[0] for r in db_users if r[1]}
        db_users_by_phone = {r[2].strip(): r[0] for r in db_users if r[2]}
        db_users_by_username = {r[3].lower().strip(): r[0] for r in db_users if r[3]}
        
        cur.execute('SELECT "CourseId", "GroupId" FROM "CourseGroups";')
        db_course_groups = set((r[0], r[1]) for r in cur.fetchall())
        
        cur.execute('SELECT "UserId", "GroupId" FROM "GroupMembers";')
        db_group_members = set((r[0], r[1]) for r in cur.fetchall())
        
        cur.execute('SELECT "CourseId", "UserId" FROM "CourseStudents";')
        db_course_students = set((r[0], r[1]) for r in cur.fetchall())

        # Check if TenantMemberships table exists in this database
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_name = 'TenantMemberships'
            );
        """)
        has_tenant_memberships = cur.fetchone()[0]

        db_tenant_members = set()
        if has_tenant_memberships:
            cur.execute('SELECT "UserId" FROM "TenantMemberships";')
            db_tenant_members = set(r[0] for r in cur.fetchall())

        print(f"[+] Loaded: {len(db_courses)} courses, {len(db_groups)} groups, {len(db_users)} users (TenantMemberships table exists: {has_tenant_memberships}).")

        # Calculated Work
        new_courses = [c for c in courses_to_create if c.lower().strip() not in db_courses]
        new_groups = [g for g in groups_to_create if g.lower().strip() not in db_groups]
        
        print("\n" + "=" * 50)
        print("                  PROPOSED WORKPLAN")
        print("=" * 50)
        print(f" 1. Create Courses    : {len(new_courses)} new to create, {len(courses_to_create) - len(new_courses)} already exist")
        print(f" 2. Create Groups     : {len(new_groups)} new to create, {len(groups_to_create) - len(new_groups)} already exist")
        print(f" 3. Define Course-Group Mappings (CourseGroups)")
        print(f" 4. Create Students   : {len(all_students)} total to evaluate")
        print(f" 5. Assign Students to Groups (GroupMembers)")
        print(f" 6. Assign Students directly to Courses (CourseStudents)")
        print("=" * 50)

        if not args.execute:
            print("\n[*] RUNNING IN DRY-RUN (SIMULATION) MODE. No database writes will be executed.")
            print("[*] To perform the actual import, run with: python import_omurhoca.py --execute")
            cur.close()
            conn.close()
            sys.exit(0)

        print("\n[*] Starting transaction...")

        # 1. Create Courses
        print("[1/6] Writing Courses...")
        course_map = {c.lower().strip(): db_courses[c.lower().strip()] for c in courses_to_create if c.lower().strip() in db_courses}
        courses_created = 0
        for c_name in new_courses:
            c_id = str(uuid.uuid4())
            cur.execute(
                'INSERT INTO "Courses" ("Id", "Title", "IsDeleted", "IsPublished", "CourseType", "Mode", "Order", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s, %s)',
                (c_id, c_name, False, True, 'Online', 'Offline', 0, datetime.datetime.utcnow())
            )
            course_map[c_name.lower().strip()] = c_id
            courses_created += 1
        print(f"      -> Created: {courses_created} courses.")

        # 2. Create Groups
        print("[2/6] Writing Groups...")
        group_map = {g.lower().strip(): db_groups[g.lower().strip()] for g in groups_to_create if g.lower().strip() in db_groups}
        groups_created = 0
        for g_name in new_groups:
            g_id = str(uuid.uuid4())
            cur.execute(
                'INSERT INTO "Groups" ("Id", "Name", "ParentId", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s, %s)',
                (g_id, g_name, None, False, datetime.datetime.utcnow())
            )
            group_map[g_name.lower().strip()] = g_id
            groups_created += 1
        print(f"      -> Created: {groups_created} groups.")

        # 3. Write CourseGroups mappings
        print("[3/6] Mapping Courses to Groups...")
        cg_created = 0
        for g_name, c_name, is_online in course_groups_to_create:
            g_id = group_map.get(g_name.lower().strip())
            c_id = course_map.get(c_name.lower().strip())
            if g_id and c_id:
                if (c_id, g_id) not in db_course_groups:
                    cg_mode = 'Online' if is_online else 'Offline'
                    cur.execute(
                        'INSERT INTO "CourseGroups" ("Id", "CourseId", "GroupId", "Mode", "AssignedAt") VALUES (%s, %s, %s, %s, %s)',
                        (str(uuid.uuid4()), c_id, g_id, cg_mode, datetime.datetime.utcnow())
                    )
                    db_course_groups.add((c_id, g_id))
                    cg_created += 1
        print(f"      -> Created: {cg_created} course-group connections.")

        # 4. Create Students (Users & TenantMemberships)
        print("[4/6] Creating Student accounts...")
        user_map = {} # norm_name -> user_uuid
        students_created = 0
        students_skipped = 0
        tm_created = 0

        for norm_name, stud in all_students.items():
            name = stud['name']
            phone = stud['phone']
            email = stud['email']
            first, last = parse_name(name)

            # Build credentials
            clean_name = normalize(name)
            if not clean_name:
                clean_name = "student_" + str(uuid.uuid4())[:8]

            phone_val = phone if phone else clean_name
            email_val = email if email else f"{clean_name}@muro.com"

            # Check duplicate
            u_id = None
            if email_val.lower().strip() in db_users_by_email:
                u_id = db_users_by_email[email_val.lower().strip()]
            elif phone_val in db_users_by_phone:
                u_id = db_users_by_phone[phone_val]
            elif phone_val in db_users_by_username:
                u_id = db_users_by_username[phone_val]

            if u_id:
                user_map[norm_name] = u_id
                students_skipped += 1
            else:
                u_id = str(uuid.uuid4())
                user_map[norm_name] = u_id
                cur.execute(
                    'INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Username", "Phone", "PasswordHash", "Role", "StudentType", "IsActive", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                    (u_id, first, last, email_val, phone_val, phone_val, email_val, 'Student', 'Active', True, False, datetime.datetime.utcnow())
                )
                db_users_by_email[email_val.lower().strip()] = u_id
                db_users_by_phone[phone_val] = u_id
                db_users_by_username[phone_val] = u_id
                students_created += 1

            # Ensure Tenant Membership exists
            if has_tenant_memberships and u_id not in db_tenant_members:
                cur.execute(
                    'INSERT INTO "TenantMemberships" ("Id", "UserId", "Role", "Status", "JoinedAt") VALUES (%s, %s, %s, %s, %s)',
                    (str(uuid.uuid4()), u_id, 'Student', 'active', datetime.datetime.utcnow())
                )
                db_tenant_members.add(u_id)
                tm_created += 1

        print(f"      -> Created: {students_created} students, Associated Tenant: {tm_created}, Skipped: {students_skipped}.")

        # 5. Assign Students to Groups (GroupMembers)
        print("[5/6] Mapping Students to Groups...")
        gm_created = 0
        gm_skipped = 0
        for norm_name, stud in all_students.items():
            u_id = user_map.get(norm_name)
            if not u_id:
                continue
            for g_name in stud['groups']:
                g_id = group_map.get(g_name.lower().strip())
                if g_id:
                    if (u_id, g_id) not in db_group_members:
                        cur.execute(
                            'INSERT INTO "GroupMembers" ("Id", "UserId", "GroupId", "Role", "Status", "AddedAt") VALUES (%s, %s, %s, %s, %s, %s)',
                            (str(uuid.uuid4()), u_id, g_id, 2, 'active', datetime.datetime.utcnow())
                        )
                        db_group_members.add((u_id, g_id))
                        gm_created += 1
                    else:
                        gm_skipped += 1
        print(f"      -> Created: {gm_created} memberships, Skipped: {gm_skipped}.")

        # 6. Assign Students directly to Courses (CourseStudents)
        print("[6/6] Mapping Students directly to Courses...")
        cs_created = 0
        cs_skipped = 0
        for norm_name, stud in all_students.items():
            u_id = user_map.get(norm_name)
            if not u_id:
                continue
            for c_name in stud['courses']:
                c_id = course_map.get(c_name.lower().strip())
                if c_id:
                    if (c_id, u_id) not in db_course_students:
                        cur.execute(
                            'INSERT INTO "CourseStudents" ("Id", "CourseId", "UserId", "AssignedAt", "ExpiresAt") VALUES (%s, %s, %s, %s, %s)',
                            (str(uuid.uuid4()), c_id, u_id, datetime.datetime.utcnow(), None)
                        )
                        db_course_students.add((c_id, u_id))
                        cs_created += 1
                    else:
                        cs_skipped += 1
        print(f"      -> Created: {cs_created} individual course assignments, Skipped: {cs_skipped}.")

        # Commit Transaction
        print("\n[*] Committing database transaction...")
        conn.commit()
        print("[+] SUCCESS: Data migration completed and committed successfully!")

    except Exception as db_err:
        print(f"\n[!] Database Transaction Error: {db_err}")
        print("[*] Rolling back changes to ensure data integrity...")
        conn.rollback()
        print("[*] Rollback completed. No changes saved.")
        sys.exit(1)
    finally:
        cur.close()
        conn.close()

if __name__ == '__main__':
    main()
