import openpyxl
import csv
import tkinter as tk
from tkinter import filedialog
import os

print("--- Mevzuat Adam %100 Risksiz Excel Ayıklayıcı ---")

# 1. Mevzuat Adam Excel'ini oku
excel_path = r"C:\Users\Rüstem\Desktop\okideolar\öğrenci listeleer\mevzuatadam.xlsx"
if not os.path.exists(excel_path):
    print(f"HATA: {excel_path} bulunamadı!")
    input("Çıkmak için Enter'a basın...")
    exit()

print("1. Mevzuat Adam excelinden ders isimleri çekiliyor...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

mevzuat_dersleri = set()
for i in range(2, ws.max_row + 1):
    ders_adi = ws.cell(row=i, column=3).value
    if ders_adi:
        # Birebir eşleşmesi için başındaki sonundaki boşlukları siliyoruz
        mevzuat_dersleri.add(str(ders_adi).strip().lower())

print(f"Başarılı! Toplam {len(mevzuat_dersleri)} adet ÖZEL ders tespit edildi.")

# 2. s7'nin CSV'sini seç
print("\n2. Lütfen s7_videolar_analiz.csv (veya s7.csv) dosyasını seçin...")
root = tk.Tk()
root.withdraw()
root.attributes('-topmost', True)
csv_path = filedialog.askopenfilename(
    title="s7'den alınan CSV dosyasını seç (s7_videolar_analiz.csv)",
    filetypes=[("CSV Dosyaları", "*.csv")]
)

if not csv_path:
    print("Dosya seçilmedi, çıkılıyor.")
    input("Çıkmak için Enter'a basın...")
    exit()

# 3. CSV'yi oku ve güncelle
output_path = os.path.join(os.path.dirname(csv_path), "kesin_s7_mevzuat_filtreli.csv")

eslesen_sayisi = 0
results = []
try:
    with open(csv_path, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        fieldnames = reader.fieldnames
        for row in reader:
            # CSV'deki Ders Adi
            csv_ders_adi = str(row.get("Ders Adi", "")).strip().lower()
            
            # Eğer bu ders adı, Mevzuat Adam listesinde BİREBİR varsa:
            if csv_ders_adi in mevzuat_dersleri:
                row["Kurum"] = "Mevzuat Adam"
                eslesen_sayisi += 1
            
            results.append(row)
            
    # Yeni dosyayı yaz
    with open(output_path, mode="w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(results)
        
    print(f"\n[+] HARİKA! İşlem tamam.")
    print(f"[+] S7'deki 5000 video içinden {eslesen_sayisi} tanesi %100 kesinlikle Mevzuat Adam olarak işaretlendi.")
    print(f"[+] Dosya şuraya kaydedildi:\n -> {output_path}")
    print("\nLÜTFEN ŞİMDİ BU YENİ DOSYAYI (kesin_s7_mevzuat_filtreli.csv) FİLEZİLLA İLE S7'YE AT VE ADINI s7.csv YAP.")
except Exception as e:
    print(f"Hata oluştu: {e}")

input("\nKapatmak için Enter'a basın...")
