import os
import sys
import uuid
import datetime
import re
import argparse
import hashlib
import psycopg2
import openpyxl
import json

def is_whitelisted(course_name):
    c = str(course_name).lower()
    if "yelda" in c: return True
    if "2025 geleneksel" in c: return True
    if "2026" in c: return True
    if "2024 eski türk edebiyatı(türkçe)" in c: return True
    if "dil bilimi soru analizi" in c and "umut hoca" in c: return True
    
    eb_courses = [
        "2024 program",
        "2024 gelişim",
        "2024 öğrenme",
        "2024 öyt",
        "2024 rehberlik",
        "2024 ölçme"
    ]
    for eb in eb_courses:
        if eb in c and "kamp" not in c: return True
    return False

def normalize(s):
    if not s:
        return ''
    s = str(s).lower().strip()
    s = s.replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
    return re.sub(r'[^a-z0-9]', '', s)

def clean_phone_number(phone):
    if not phone:
        return None
    phone_str = str(phone).strip()
    main_part = phone_str.split('.')[0].split(',')[0]
    digits = "".join(c for c in main_part if c.isdigit())
    while digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 12 and digits.startswith("90"):
        digits = digits[2:]
    return digits if digits else None

def generate_mock_phone(name):
    h = hashlib.md5(name.encode('utf-8')).hexdigest()
    val = int(h, 16) % 10000000
    return f"111{val:07d}"

def parse_name(full_name):
    if not full_name:
        return "", ""
    parts = str(full_name).strip().split()
    if len(parts) == 0:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]

def generate_mock_email(full_name):
    first, last = parse_name(full_name)
    first_norm = normalize(first)
    last_norm = normalize(last)
    if not first_norm:
        return f"student_{str(uuid.uuid4())[:8]}@akademikmasa.com"
    if not last_norm:
        return f"{first_norm}@akademikmasa.com"
    return f"{first_norm}.{last_norm}@akademikmasa.com"

def generate_password(full_name, phone):
    phone_suffix = phone[-2:] if (phone and len(phone) >= 2) else "99"
    first, last = parse_name(full_name)
    
    first_clean = normalize(first)
    if not first_clean:
        first_clean = "student"
        
    last_clean = normalize(last)
    last_char = last_clean[0] if last_clean else "a"
    
    return f"{first_clean}.{phone_suffix}.{last_char}"

def main():
    parser = argparse.ArgumentParser(description="AKADEMIK MASA LMS Importer")
    parser.add_argument("--host", help="PostgreSQL host")
    parser.add_argument("--port", type=int, help="PostgreSQL port")
    parser.add_argument("--dbname", help="PostgreSQL database name")
    parser.add_argument("--user", help="PostgreSQL username")
    parser.add_argument("--password", help="PostgreSQL password")
    parser.add_argument("--execute", action="store_true", help="Execute the database writes (defaults to dry-run)")
    args = parser.parse_args()

    print("=" * 60)
    print("      AKADEMIK MASA LMS - COURSES, GROUPS & STUDENTS IMPORT")
    print("=" * 60)

    conn_params = {}
    paths = [
        "src/MURO.API/appsettings.json",
        "appsettings.json",
        "../src/MURO.API/appsettings.json"
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    content = f.read()
                    match = re.search(r'"DefaultConnection"\s*:\s*"([^"]+)"', content)
                    if match:
                        conn_str = match.group(1)
                        for item in conn_str.split(';'):
                            if '=' in item:
                                k, v = item.split('=', 1)
                                k_lower = k.strip().lower()
                                v_val = v.strip()
                                if k_lower == 'host': conn_params['host'] = v_val
                                elif k_lower == 'port': conn_params['port'] = int(v_val)
                                elif k_lower == 'database': conn_params['dbname'] = v_val
                                elif k_lower == 'username': conn_params['user'] = v_val
                                elif k_lower == 'password': conn_params['password'] = v_val
            except Exception:
                pass

    if args.host: conn_params['host'] = args.host
    if args.port: conn_params['port'] = args.port
    if args.dbname: conn_params['dbname'] = args.dbname
    if args.user: conn_params['user'] = args.user
    if args.password: conn_params['password'] = args.password

    # Try reading from .env files
    env_paths = [".env", "../.env", "../../.env"]
    for ep in env_paths:
        if os.path.exists(ep):
            try:
                with open(ep, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith("#") and "=" in line:
                            k, v = line.split("=", 1)
                            k = k.strip()
                            v = v.strip()
                            if k == "DB_USER": conn_params["user"] = v
                            elif k == "DB_PASSWORD": conn_params["password"] = v
                            elif k == "DB_NAME": conn_params["dbname"] = v
            except Exception:
                pass

    # Try resolving docker container IP just in case localhost is mapped to another Postgres
    try:
        import subprocess
        result = subprocess.check_output(['docker', 'inspect', '-f', '{{range .NetworkSettings.Networks}}{{.IPAddress}}{{end}}', 'muro_akm_postgres'], stderr=subprocess.DEVNULL)
        ip = result.decode('utf-8').strip()
        if ip:
            conn_params["host"] = ip
    except Exception:
        pass

    if 'host' not in conn_params: conn_params['host'] = 'localhost'
    if 'port' not in conn_params: conn_params['port'] = 5432
    if 'dbname' not in conn_params: conn_params['dbname'] = 'muro_demo'
    if 'user' not in conn_params: conn_params['user'] = 'muro_user'
    if 'password' not in conn_params: conn_params['password'] = 'MuroDem0_2026!Str0ng'

    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, "akademikmasa.xlsx")

    # LOAD JSON FILES FOR SESSION MAPPING AND FILTERING
    akm_isim_path = os.path.join(script_dir, "akm_isim_eslestirme.json")
    akm_guncel_path = os.path.join(script_dir, "akm_guncel_kayitlar.json")
    
    link_to_name = {}
    if os.path.exists(akm_isim_path):
        with open(akm_isim_path, "r", encoding="utf-8") as f:
            for item in json.load(f):
                if item.get("classroomLink") and item.get("courseName"):
                    link_to_name[item["classroomLink"]] = item["courseName"].strip()
    else:
        print(f"[!] Warning: {akm_isim_path} not found!")
                    
    course_to_recordings = {}
    if os.path.exists(akm_guncel_path):
        with open(akm_guncel_path, "r", encoding="utf-8") as f:
            for item in json.load(f):
                link = item.get("classroomLink")
                vcount = item.get("videoCount", 0)
                recs = item.get("recordings", [])
                
                real_name = link_to_name.get(link)
                if not real_name:
                    real_name = item.get("courseName")
                    
                if real_name and vcount > 0 and len(recs) > 0:
                    course_to_recordings[real_name.lower().strip()] = recs
    else:
        print(f"[!] Warning: {akm_guncel_path} not found!")

    print(f"[*] Loaded JSON Mapping: Found {len(course_to_recordings)} valid courses with > 0 videos.")

    print(f"[*] Reading '{file_path}'...")
    if not os.path.exists(file_path):
        print(f"[!] Error: File '{file_path}' not found.")
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    s1 = wb['Sayfa1']
    rows1 = list(s1.iter_rows(values_only=True))
    print(f"    - Sayfa1 rows: {len(rows1)}")
    
    all_students = {}
    for r in rows1[2:]:
        name = r[1]
        phone = r[2]
        email = r[3]
        
        if name:
            name_str = str(name).strip()
            norm_name = normalize(name_str)
            
            cleaned_phone = clean_phone_number(phone)
            is_mock_phone = False
            if not cleaned_phone:
                cleaned_phone = generate_mock_phone(name_str)
                is_mock_phone = True
                
            cleaned_email = str(email).strip().lower() if email else None
            if not cleaned_email or "@" not in cleaned_email:
                cleaned_email = generate_mock_email(name_str)
                
            pwd = generate_password(name_str, cleaned_phone)
            
            all_students[norm_name] = {
                'name': name_str,
                'phone': cleaned_phone,
                'email': cleaned_email,
                'password': pwd,
                'is_mock_phone': is_mock_phone,
                'groups': set()
            }
    print(f"    - Loaded students from Sayfa1: {len(all_students)} students mapped.")

    s2 = wb['Sayfa2']
    rows2 = list(s2.iter_rows(values_only=True))
    print(f"    - Sayfa2 rows: {len(rows2)}")

    courses_to_create = set()
    groups_to_create = set()
    course_groups_to_create = [] 
    groups_metadata = {}

    for r in rows2[2:]:
        g_id = r[1]
        g_name = r[2]
        c_name = r[3]
        mode = r[5]
        if g_id and g_name:
            g_id_str = str(g_id).strip()
            g_name_str = str(g_name).strip()
            groups_to_create.add(g_id_str)
            groups_metadata[g_id_str] = g_name_str
            
            if c_name and str(c_name).strip() != "Gösterilecek veri yok.":
                c_name_str = str(c_name).strip()
                
                # Apply Filters
                if not is_whitelisted(c_name_str):
                    continue
                if c_name_str.lower().strip() not in course_to_recordings:
                    continue

                courses_to_create.add(c_name_str)
                is_online = (str(mode).strip().lower() == "canlı")
                course_groups_to_create.append((g_id_str, c_name_str, is_online))

    for r in rows2[2:]:
        g_id = r[10]
        g_name = r[11]
        ad_soyad = r[12]
        if g_id and g_name and ad_soyad and "gösterilecek" not in str(ad_soyad).lower():
            g_id_str = str(g_id).strip()
            g_name_str = str(g_name).strip()
            student_name_str = str(ad_soyad).strip()
            norm_name = normalize(student_name_str)

            groups_to_create.add(g_id_str)
            groups_metadata[g_id_str] = g_name_str

            if norm_name not in all_students:
                cleaned_phone = generate_mock_phone(student_name_str)
                cleaned_email = generate_mock_email(student_name_str)
                pwd = generate_password(student_name_str, cleaned_phone)
                all_students[norm_name] = {
                    'name': student_name_str,
                    'phone': cleaned_phone,
                    'email': cleaned_email,
                    'password': pwd,
                    'is_mock_phone': True,
                    'groups': set()
                }
            all_students[norm_name]['groups'].add(g_id_str)

    hierarchy_path = os.path.join(script_dir, "akm_grup_hiyerarsi.json")
    node_parents = {} 
    
    if os.path.exists(hierarchy_path):
        try:
            with open(hierarchy_path, 'r', encoding='utf-8') as f:
                h_nodes = json.load(f)
            
            last_seen_id_by_name = {}
            for node in h_nodes:
                n_id = str(node.get('id', '')).strip()
                n_name = str(node.get('name', '')).strip()
                parent_name = str(node.get('parent', '')).strip()
                
                if not n_id or not n_name:
                    continue
                
                groups_metadata[n_id] = n_name
                groups_to_create.add(n_id)
                last_seen_id_by_name[n_name.lower().strip()] = n_id
                
                if parent_name:
                    p_id = last_seen_id_by_name.get(parent_name.lower().strip())
                    if p_id and p_id != n_id:
                        node_parents[n_id] = p_id
            print(f"    - Loaded hierarchy from JSON: {len(node_parents)} parent-child links resolved.")
        except Exception as e:
            print(f"    - Warning: Failed to load group hierarchy JSON: {e}")

    print(f"\n[+] Parsed Excel & JSON Data successfully:")
    print(f"    - Unique Groups to evaluate: {len(groups_to_create)}")
    print(f"    - Unique Courses to create : {len(courses_to_create)} (AFTER FILTERING)")
    print(f"    - Unique Students in total : {len(all_students)}")
    print(f"    - Course-Group mappings    : {len(course_groups_to_create)}")

    print(f"\n[*] Connecting to PostgreSQL database...")
    try:
        conn = psycopg2.connect(**conn_params)
        cur = conn.cursor()
        print("[+] Connected successfully!")
    except Exception as e:
        print(f"[!] Database Connection Failed: {e}")
        sys.exit(1)

    try:
        print("[*] Retrieving existing database records...")
        
        cur.execute('SELECT "Id", "Title" FROM "Courses" WHERE "IsDeleted" = False;')
        db_courses = {r[1].lower().strip(): str(r[0]) for r in cur.fetchall()}
        
        cur.execute('SELECT "Id", "Name", "Description", "ParentId" FROM "Groups" WHERE "IsDeleted" = False;')
        db_groups_raw = cur.fetchall()
        db_groups_by_okina_id = {}
        for g_id, g_name, g_desc, g_parent in db_groups_raw:
            match = re.search(r'Okina Group ID:\s*(\d+)', str(g_desc))
            if match:
                okina_id = match.group(1)
                db_groups_by_okina_id[okina_id] = str(g_id)
        
        cur.execute('SELECT "Id", "Email", "Phone", "Username" FROM "Users";')
        db_users = cur.fetchall()
        db_users_by_email = {str(r[1]).lower().strip(): str(r[0]) for r in db_users if r[1]}
        db_users_by_phone = {str(r[2]).strip(): str(r[0]) for r in db_users if r[2]}
        db_users_by_username = {str(r[3]).lower().strip(): str(r[0]) for r in db_users if r[3]}
        
        cur.execute('SELECT "CourseId", "GroupId" FROM "CourseGroups";')
        db_course_groups = set((str(r[0]), str(r[1])) for r in cur.fetchall())
        
        cur.execute('SELECT "UserId", "GroupId" FROM "GroupMembers";')
        db_group_members = set((str(r[0]), str(r[1])) for r in cur.fetchall())
        
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_name = 'TenantMemberships'
            );
        """)
        has_tenant_memberships = cur.fetchone()[0]

        db_tenant_members = set()
        if has_tenant_memberships:
            cur.execute('SELECT "UserId" FROM "TenantMemberships";')
            db_tenant_members = set(str(r[0]) for r in cur.fetchall())

        print(f"[+] Loaded: {len(db_courses)} courses, {len(db_groups_by_okina_id)} matched groups, {len(db_users)} users.")

        new_courses = [c for c in courses_to_create if c.lower().strip() not in db_courses]
        new_groups_okina_ids = [g_id for g_id in groups_to_create if g_id not in db_groups_by_okina_id]
        
        print("\n" + "=" * 50)
        print("                  PROPOSED WORKPLAN")
        print("=" * 50)
        print(f" 1. Create Courses    : {len(new_courses)} new to create, {len(courses_to_create) - len(new_courses)} already exist")
        print(f" 2. Create Groups     : {len(new_groups_okina_ids)} new to create, {len(groups_to_create) - len(new_groups_okina_ids)} already exist")
        print(f" 3. Define Course-Group Mappings")
        print(f" 4. Create Students   : {len(all_students)} total to evaluate")
        print(f" 5. Assign Students to Groups")
        print(f" 6. Assign Sessions   : Will process {len(courses_to_create)} valid courses.")
        print("=" * 50)

        if not args.execute:
            print("\n[*] RUNNING IN DRY-RUN (SIMULATION) MODE. No database writes will be executed.")
            print("[*] To perform the actual import, run with: python import_akm.py --execute")
            cur.close()
            conn.close()
            sys.exit(0)

        print("\n[*] Starting transaction...")

        print("[1/6] Writing Courses...")
        course_map = {c.lower().strip(): db_courses[c.lower().strip()] for c in courses_to_create if c.lower().strip() in db_courses}
        courses_created = 0
        for c_name in new_courses:
            c_id = str(uuid.uuid4())
            cur.execute(
                'INSERT INTO "Courses" ("Id", "Title", "IsDeleted", "IsPublished", "CourseType", "Mode", "Order", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s, %s)',
                (c_id, c_name, False, True, 'Online', 'Offline', 0, datetime.datetime.utcnow())
            )
            course_map[c_name.lower().strip()] = c_id
            courses_created += 1
        print(f"      -> Created: {courses_created} courses.")

        print("[2/6] Writing Groups...")
        group_uuid_map = {okina_id: db_uuid for okina_id, db_uuid in db_groups_by_okina_id.items()}
        groups_created = [0] 

        def get_or_create_group(okina_id):
            if okina_id in group_uuid_map:
                return group_uuid_map[okina_id]
                
            g_name = groups_metadata.get(okina_id, f"Group {okina_id}")
            parent_okina_id = node_parents.get(okina_id)
            
            parent_uuid = None
            if parent_okina_id:
                parent_uuid = get_or_create_group(parent_okina_id)
                
            db_uuid = str(uuid.uuid4())
            desc = f"Okina Group ID: {okina_id}"
            cur.execute(
                'INSERT INTO "Groups" ("Id", "Name", "Description", "ParentId", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s)',
                (db_uuid, g_name, desc, parent_uuid, False, datetime.datetime.utcnow())
            )
            group_uuid_map[okina_id] = db_uuid
            db_groups_by_okina_id[okina_id] = db_uuid
            groups_created[0] += 1
            return db_uuid

        for g_id in groups_to_create:
            get_or_create_group(g_id)
            
        print(f"      -> Created: {groups_created[0]} groups.")

        print("[3/6] Mapping Courses to Groups...")
        cg_created = 0
        for g_id_str, c_name, is_online in course_groups_to_create:
            g_uuid = group_uuid_map.get(g_id_str)
            c_uuid = course_map.get(c_name.lower().strip())
            if g_uuid and c_uuid:
                if (c_uuid, g_uuid) not in db_course_groups:
                    cg_mode = 'Online' if is_online else 'Offline'
                    cur.execute(
                        'INSERT INTO "CourseGroups" ("Id", "CourseId", "GroupId", "Mode", "AssignedAt") VALUES (%s, %s, %s, %s, %s)',
                        (str(uuid.uuid4()), c_uuid, g_uuid, cg_mode, datetime.datetime.utcnow())
                    )
                    db_course_groups.add((c_uuid, g_uuid))
                    cg_created += 1
        print(f"      -> Created: {cg_created} course-group connections.")

        print("[4/6] Creating Student accounts...")
        user_map = {}
        students_created = 0
        students_skipped = 0
        tm_created = 0

        for norm_name, stud in all_students.items():
            name = stud['name']
            phone = stud['phone']
            email = stud['email']
            pwd = stud['password']
            first, last = parse_name(name)

            phone_val = phone
            email_val = email

            u_id = None
            if email_val.lower().strip() in db_users_by_email:
                u_id = db_users_by_email[email_val.lower().strip()]
            elif phone_val in db_users_by_phone:
                u_id = db_users_by_phone[phone_val]
            elif phone_val in db_users_by_username:
                u_id = db_users_by_username[phone_val]

            if u_id:
                user_map[norm_name] = u_id
                students_skipped += 1
            else:
                u_id = str(uuid.uuid4())
                user_map[norm_name] = u_id
                cur.execute(
                    'INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Username", "Phone", "PasswordHash", "Role", "StudentType", "IsActive", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                    (u_id, first, last, email_val, phone_val, phone_val, pwd, 'Student', 'Active', True, False, datetime.datetime.utcnow())
                )
                db_users_by_email[email_val.lower().strip()] = u_id
                db_users_by_phone[phone_val] = u_id
                db_users_by_username[phone_val] = u_id
                students_created += 1

            if has_tenant_memberships and u_id not in db_tenant_members:
                cur.execute(
                    'INSERT INTO "TenantMemberships" ("Id", "UserId", "Role", "Status", "JoinedAt") VALUES (%s, %s, %s, %s, %s)',
                    (str(uuid.uuid4()), u_id, 'Student', 'active', datetime.datetime.utcnow())
                )
                db_tenant_members.add(u_id)
                tm_created += 1

        print(f"      -> Created: {students_created} students, Associated Tenant: {tm_created}, Skipped: {students_skipped}.")

        print("[5/6] Mapping Students to Groups...")
        gm_created = 0
        gm_skipped = 0
        for norm_name, stud in all_students.items():
            u_id = user_map.get(norm_name)
            if not u_id:
                continue
            for g_id_str in stud['groups']:
                g_uuid = group_uuid_map.get(g_id_str)
                if g_uuid:
                    if (u_id, g_uuid) not in db_group_members:
                        cur.execute(
                            'INSERT INTO "GroupMembers" ("Id", "UserId", "GroupId", "Role", "Status", "AddedAt") VALUES (%s, %s, %s, %s, %s, %s)',
                            (str(uuid.uuid4()), u_id, g_uuid, 2, 'active', datetime.datetime.utcnow())
                        )
                        db_group_members.add((u_id, g_uuid))
                        gm_created += 1
                    else:
                        gm_skipped += 1
        print(f"      -> Created: {gm_created} memberships, Skipped: {gm_skipped}.")

        print("[6/6] Writing BBB Sessions...")
        sessions_created = 0
        sessions_skipped = 0
        
        cur.execute('SELECT "CourseId", "BbbMeetingId" FROM "Sessions" WHERE "IsDeleted" = False;')
        db_sessions = set((str(r[0]), str(r[1])) for r in cur.fetchall())
        
        for c_name in courses_to_create:
            c_uuid = course_map.get(c_name.lower().strip())
            recs = course_to_recordings.get(c_name.lower().strip(), [])
            if c_uuid and recs:
                for idx, rec in enumerate(recs):
                    rec_id = rec.get("recordID")
                    v_name = rec.get("videoName", f"Kayıt {idx+1}")
                    
                    if rec_id:
                        if (c_uuid, rec_id) not in db_sessions:
                            s_id = str(uuid.uuid4())
                            cur.execute(
                                'INSERT INTO "Sessions" ("Id", "CourseId", "Title", "BbbMeetingId", "IsDeleted", "CreatedAt", "Status", "Order", "RecordingEnabled", "IsFree") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                                (s_id, c_uuid, v_name, rec_id, False, datetime.datetime.utcnow(), 3, idx, True, False)
                            )
                            db_sessions.add((c_uuid, rec_id))
                            sessions_created += 1
                        else:
                            sessions_skipped += 1
        print(f"      -> Created: {sessions_created} sessions, Skipped: {sessions_skipped}.")

        print("\n[*] Committing database transaction...")
        conn.commit()
        print("[+] SUCCESS: Data migration completed and committed successfully!")

    except Exception as db_err:
        print(f"\n[!] Database Transaction Error: {db_err}")
        print("[*] Rolling back changes to ensure data integrity...")
        conn.rollback()
        print("[*] Rollback completed. No changes saved.")
        sys.exit(1)
    finally:
        cur.close()
        conn.close()

if __name__ == '__main__':
    main()
