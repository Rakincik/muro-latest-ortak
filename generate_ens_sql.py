import openpyxl
import re
import uuid
import datetime

excel_file = "türkçeöabtdeyiz.xlsx"
sql_output_file = "import_ens_users.sql"

print(f"Reading Excel file '{excel_file}'...")
wb = openpyxl.load_workbook(excel_file, data_only=True)
s1 = wb["Sayfa1"]
s2 = wb["Sayfa2"]

def clean_tr(text):
    tr_map = {
        'ı': 'i', 'ğ': 'g', 'ü': 'u', 'ş': 's', 'ö': 'o', 'ç': 'c',
        'İ': 'i', 'Ğ': 'g', 'Ü': 'u', 'Ş': 's', 'Ö': 'o', 'Ç': 'c',
        'I': 'i', 'Ş': 's', 'Ğ': 'g', 'Ü': 'u', 'Ö': 'o', 'Ç': 'c'
    }
    for k, v in tr_map.items():
        text = text.replace(k, v)
    return text

# Parse Sayfa1: User details
users = []
seen_phones = set()

# Start from row 3 (which is index 2, since index 0 is None, index 1 is header row)
for row in list(s1.iter_rows(values_only=True))[2:]:
    name = row[1]
    phone = row[2]
    email = row[3]

    if not name:
        continue

    name_str = str(name).strip()
    name_key = name_str.lower()

    # Process Phone
    phone_str = ""
    if phone:
        phone_str = "".join(filter(str.isdigit, str(phone)))
        if phone_str.startswith("90") and len(phone_str) > 10:
            phone_str = phone_str[2:]

    original_phone_str = phone_str
    if not phone_str:
        # Fallback if no phone (though we checked there are none, we keep it safe)
        phone_str = f"no_phone_{uuid.uuid4().hex[:6]}"

    # Deduplicate phones just in case
    counter = 1
    base_phone_str = phone_str
    while phone_str in seen_phones:
        phone_str = f"{base_phone_str}_{counter}"
        counter += 1

    seen_phones.add(phone_str)

    # If there is a real email in column D, use it. Otherwise generate one.
    if email and "@" in str(email):
        email_str = str(email).strip().lower()
    else:
        email_str = f"student_{phone_str}@muro.com"

    # Password generation logic: isim.telefonunson2_hanesi.soyismin_ilk_harfi
    parts = name_str.split()
    if len(parts) >= 2:
        isim = "".join(parts[:-1]).lower()
        soyisim = parts[-1].lower()
    else:
        isim = name_str.lower()
        soyisim = name_str.lower()

    isim = clean_tr(isim)
    soyisim_ilk_harf = clean_tr(soyisim[0]) if soyisim else "x"

    if original_phone_str and len(original_phone_str) >= 2:
        son2 = original_phone_str[-2:]
        password = f"{isim}.{son2}.{soyisim_ilk_harf}"
    else:
        password = f"{isim}.99.{soyisim_ilk_harf}"

    # First name and last name division
    first_name = " ".join(parts[:-1]) if len(parts) >= 2 else name_str
    last_name = parts[-1] if len(parts) >= 2 else "-"

    users.append({
        'name_key': name_key,
        'name': name_str,
        'first_name': first_name,
        'last_name': last_name,
        'phone': phone_str,
        'actual_phone': original_phone_str,
        'email': email_str,
        'password': password
    })

# Parse Sayfa2: Groups, Courses, and Mappings
# Left side starts at row index 2 (row 3)
rows2 = list(s2.iter_rows(values_only=True))

unique_groups = set()
unique_courses = set()
group_course_mappings = [] # list of (group_name, course_title, mode)
student_group_mappings = [] # list of (student_name_key, group_name)

# Header is on index 1
header_row = rows2[1]

# We need to collect groups and courses from Left table
for row in rows2[2:]:
    g_id = row[1]
    g_name = row[2]
    c_title = row[3]
    c_mode = row[5]

    if g_name:
        unique_groups.add(str(g_name).strip())
    if c_title and str(c_title).strip() != "Gösterilecek veri yok.":
        unique_courses.add(str(c_title).strip())
        if g_name:
            group_course_mappings.append((str(g_name).strip(), str(c_title).strip(), str(c_mode).strip()))

# Right table: Group members
for row in rows2[2:]:
    g_name_r = row[11]
    ad_soyad = row[12]
    rol = row[13]

    if g_name_r:
        unique_groups.add(str(g_name_r).strip())
    if ad_soyad and str(ad_soyad).strip() != "Gösterilecek veri yok.":
        student_group_mappings.append((str(ad_soyad).strip().lower(), str(g_name_r).strip()))

# Generate SQL script
sql_lines = [
    "-- ==================================================",
    "-- MURO LMS - ENS Tenant Restoration Script",
    f"-- Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    "-- ==================================================",
    "BEGIN;",
    ""
]

# 1. GROUPS
sql_lines.append("-- 1. CREATE GROUPS")
group_uuids = {}
for g in sorted(unique_groups):
    g_id = str(uuid.uuid4())
    group_uuids[g] = g_id
    safe_g = g.replace("'", "''")
    sql_lines.append(f"""
INSERT INTO "Groups" ("Id", "Name", "IsDeleted", "CreatedAt")
SELECT '{g_id}', '{safe_g}', false, NOW()
WHERE NOT EXISTS (SELECT 1 FROM "Groups" WHERE "Name" = '{safe_g}' AND "IsDeleted" = false);
""".strip())

sql_lines.append("")

# 2. COURSES
sql_lines.append("-- 2. CREATE COURSES")
course_uuids = {}
for c in sorted(unique_courses):
    c_id = str(uuid.uuid4())
    course_uuids[c] = c_id
    safe_c = c.replace("'", "''")
    sql_lines.append(f"""
INSERT INTO "Courses" ("Id", "Title", "IsDeleted", "IsPublished", "CourseType", "Mode", "Order", "CreatedAt")
SELECT '{c_id}', '{safe_c}', false, true, 'Online', 'Offline', 0, NOW()
WHERE NOT EXISTS (SELECT 1 FROM "Courses" WHERE "Title" = '{safe_c}' AND "IsDeleted" = false);
""".strip())

sql_lines.append("")

# 3. COURSE GROUPS (MAPPINGS)
sql_lines.append("-- 3. COURSE-GROUP MAPPINGS")
for g_name, c_title, mode in group_course_mappings:
    safe_g = g_name.replace("'", "''")
    safe_c = c_title.replace("'", "''")
    db_mode = 'Online' if mode == 'Canlı' else 'Offline'
    
    sql_lines.append(f"""
INSERT INTO "CourseGroups" ("Id", "CourseId", "GroupId", "Mode", "AssignedAt")
SELECT gen_random_uuid(), c."Id", g."Id", '{db_mode}', NOW()
FROM "Courses" c, "Groups" g
WHERE c."Title" = '{safe_c}' AND g."Name" = '{safe_g}'
  AND NOT EXISTS (
      SELECT 1 FROM "CourseGroups" cg 
      JOIN "Courses" c2 ON cg."CourseId" = c2."Id"
      JOIN "Groups" g2 ON cg."GroupId" = g2."Id"
      WHERE c2."Title" = '{safe_c}' AND g2."Name" = '{safe_g}'
  );
""".strip())

sql_lines.append("")

# 4. USERS
sql_lines.append("-- 4. CREATE USERS (STUDENTS)")
for u in users:
    safe_fn = u['first_name'].replace("'", "''")
    safe_ln = u['last_name'].replace("'", "''")
    safe_e = u['email'].replace("'", "''")
    safe_p = u['phone'].replace("'", "''")
    safe_pw = u['password'].replace("'", "''")
    
    sql_lines.append(f"""
INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Username", "Phone", "PasswordHash", "Role", "StudentType", "IsActive", "IsDeleted", "CreatedAt")
SELECT gen_random_uuid(), '{safe_fn}', '{safe_ln}', '{safe_e}', '{safe_p}', '{safe_p}', '{safe_pw}', 'Student', 'Active', true, false, NOW()
WHERE NOT EXISTS (SELECT 1 FROM "Users" WHERE "Username" = '{safe_p}');
""".strip())

sql_lines.append("")

# 5. GROUP MEMBERSHIPS
sql_lines.append("-- 5. DEFINE STUDENT-GROUP MEMBERSHIPS")
for name_key, g_name in student_group_mappings:
    # Find user
    matched_user = next((u for u in users if u['name_key'] == name_key), None)
    if not matched_user:
        continue
    
    safe_p = matched_user['phone'].replace("'", "''")
    safe_g = g_name.replace("'", "''")
    
    sql_lines.append(f"""
INSERT INTO "GroupMembers" ("Id", "GroupId", "UserId", "Role", "Status", "AddedAt")
SELECT gen_random_uuid(), g."Id", u."Id", 2, 'active', NOW()
FROM "Users" u, "Groups" g
WHERE (u."Username" = '{safe_p}' OR u."Email" = '{matched_user['email']}') AND g."Name" = '{safe_g}'
  AND NOT EXISTS (
      SELECT 1 FROM "GroupMembers" gm 
      JOIN "Users" u2 ON gm."UserId" = u2."Id"
      JOIN "Groups" g2 ON gm."GroupId" = g2."Id"
      WHERE (u2."Username" = '{safe_p}' OR u2."Email" = '{matched_user['email']}') AND g2."Name" = '{safe_g}'
  );
""".strip())

sql_lines.append("")
sql_lines.append("COMMIT;")

with open(sql_output_file, "w", encoding="utf-8") as f:
    f.write("\n".join(sql_lines))

print(f"Generated {sql_output_file} successfully.")
print(f"Summary:")
print(f"  - Unique Groups  : {len(unique_groups)}")
print(f"  - Unique Courses : {len(unique_courses)}")
print(f"  - Unique Users   : {len(users)}")
print(f"  - Course-Group Mappings: {len(group_course_mappings)}")
print(f"  - Student-Group Mappings: {len(student_group_mappings)}")
