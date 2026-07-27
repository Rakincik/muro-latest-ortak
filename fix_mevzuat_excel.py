import openpyxl
import json

# 1. JSON verisini oku (90 Ders)
json_path = r"okinar dersler/mevzuatadam.okinar.com_recordings.json"
excel_path = r"C:\Users\Rüstem\Desktop\okideolar\öğrenci listeleer\mevzuatadam.xlsx"

print(f"Reading recordings from {json_path}...")
with open(json_path, "r", encoding="utf-8") as f:
    recordings = json.load(f)
print(f"Loaded {len(recordings)} recordings.")

# 2. Otomatik Sınıflandırma Kuralları
def classify_course(course_name):
    nl = course_name.lower()
    
    # Tarım ve Orman
    if 'tob' in nl or 'tarım' in nl or 'orman' in nl:
        return 118, 'TOB Ortak Konular Soru Çözümü', 'Canlı' if 'soru' in nl else 'Video'
        
    # Gençlik ve Spor
    if any(k in nl for k in ['gsb', 'gençlik', 'spor', 'yurt', '351', '3289', '6222', '7405', 'yönerge', 'taşra']):
        is_live = 'Canlı' if 'soru' in nl else 'Video'
        if any(k in nl for k in ['bütçe', 'muhasebe', '4483', '4734', '4735', '5018', '5442', '6222', '6698', '7405']):
            return 116, 'GSB Alan Soru Çözümü', is_live
        return 115, 'GSB Ortak Konular Soru Çözümü', is_live
        
    # Adalet / CTE
    if any(k in nl for k in ['adalet', 'ceza', 'infaz', 'yargıtay', 'uyap', 'segbis', 'adliye', 'başsavcılık', 'yazı işleri', 'bim', 'idare mahkemesi', 'vergi mahkemesi', 'hakim', 'savcı', 'hmk', 'cmk', 'tck', 'tebligat', '7201', '5275', '2802', '5235', '5271', 'harçlar', '492', 'meden']):
        is_live = 'Canlı' if 'soru' in nl else 'Video'
        if any(k in nl for k in ['infaz', 'tevkif', 'ceza ve güvenlik']):
            return 110, 'Ceza ve Tevkifevleri', is_live
        return 109, 'Adalet Bakanlığı', is_live
        
    # İçişleri
    if any(k in nl for k in ['belediye', 'il özel', 'il idare', '5393', '5302', '5442']):
        return 111, 'İçişleri Bakanlığı', 'Canlı' if 'soru' in nl else 'Video'
        
    # MEB
    if any(k in nl for k in ['meb', 'türkçe', 'dil bilgisi', 'rehberlik']):
        return 113, 'MEB', 'Canlı' if 'soru' in nl else 'Video'
        
    # Genel / Ortak (Mevzuat Adam Grubu)
    return 108, 'Mevzuat Adam', 'Video'

# 3. Excel dosyasını yükle
print(f"Loading Excel file from {excel_path}...")
wb = openpyxl.load_workbook(excel_path)
ws = wb['Sayfa2']

# Sağ taraftaki öğrenci eşleştirmelerini (K'dan N'ye) belleğe alalım ki kaybolmasınlar
student_rows = []
for row in range(2, ws.max_row + 1):
    student_rows.append([ws.cell(row=row, column=col).value for col in range(11, 15)]) # K to N
print(f"Retained {len(student_rows)} student mapping rows from Excel.")

# Sol taraftaki eski ders atamalarını temizleyelim (A'dan E'ye)
max_clean = max(ws.max_row, 150)
for row in range(2, max_clean + 1):
    for col in range(1, 6): # A to E
        ws.cell(row=row, column=col, value=None)

# 4. 90 dersi yeni satırlar olarak yazalım ve sağdaki öğrencileri yanlarına yerleştirelim
for i, item in enumerate(recordings):
    row_idx = i + 2
    c_name = item['ders']
    c_desc = item['aciklama']
    
    gid, gname, mod = classify_course(c_name)
    
    # Sol taraf (Dersler)
    ws.cell(row=row_idx, column=1, value=gid)
    ws.cell(row=row_idx, column=2, value=gname)
    ws.cell(row=row_idx, column=3, value=c_name)
    ws.cell(row=row_idx, column=4, value=c_desc)
    ws.cell(row=row_idx, column=5, value=mod)
    
    # Sağ taraf (Öğrenciler - eğer o satırda öğrenci kalmışsa yerleştir)
    if i < len(student_rows):
        for col_offset, val in enumerate(student_rows[i]):
            ws.cell(row=row_idx, column=11 + col_offset, value=val)
    else:
        # 41 satırdan sonrası için öğrenci kısımlarını temizle
        for col_offset in range(4):
            ws.cell(row=row_idx, column=11 + col_offset, value=None)

wb.save(excel_path)
print("Excel file successfully updated with all 90 courses mapped to groups!")
