import sys
import os
import argparse
import uuid
import datetime
import re
import json
import openpyxl
import psycopg2

def clean_phone_number(phone):
    if not phone:
        return None
    phone_str = str(phone).strip()
    # Handle scientific notation or decimal parts from Excel (e.g. "5069993427.0")
    main_part = phone_str.split('.')[0].split(',')[0]
    # Keep only digits
    digits = "".join(c for c in main_part if c.isdigit())
    # Strip leading zeros
    while digits.startswith("0"):
        digits = digits[1:]
    # If 12 digits starting with 90, strip 90
    if len(digits) == 12 and digits.startswith("90"):
        digits = digits[2:]
    return digits if digits else None

def parse_name(full_name):
    if not full_name:
        return "", ""
    parts = full_name.strip().split()
    if len(parts) == 0:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]

def discover_connection_string():
    paths = [
        "src/MURO.API/appsettings.json",
        "appsettings.json",
        "../src/MURO.API/appsettings.json"
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    content = f.read()
                    match = re.search(r'"DefaultConnection"\s*:\s*"([^"]+)"', content)
                    if match:
                        conn_str = match.group(1)
                        params = {}
                        for item in conn_str.split(';'):
                            if '=' in item:
                                k, v = item.split('=', 1)
                                k_lower = k.strip().lower()
                                v_val = v.strip()
                                if k_lower == 'host': params['host'] = v_val
                                elif k_lower == 'port': params['port'] = int(v_val)
                                elif k_lower == 'database': params['dbname'] = v_val
                                elif k_lower == 'username': params['user'] = v_val
                                elif k_lower == 'password': params['password'] = v_val
                        if params:
                            return params, p
            except Exception:
                pass
    return None, None

def main():
    parser = argparse.ArgumentParser(description="MURO Excel Importer - Dereceuzem")
    parser.add_argument("--host", help="PostgreSQL host")
    parser.add_argument("--port", type=int, help="PostgreSQL port")
    parser.add_argument("--dbname", help="PostgreSQL database name")
    parser.add_argument("--user", help="PostgreSQL username")
    parser.add_argument("--password", help="PostgreSQL password")
    parser.add_argument("--no-db", action="store_true", help="Skip db connection and only parse Excel data")
    parser.add_argument("--execute", action="store_true", help="Execute inserts (default is dry-run)")
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("            MURO LMS - DERECEUZEM DATA IMPORTER")
    print("=" * 60)
    
    # Connection parameters discovery
    conn_params = {}
    disc_params, config_path = discover_connection_string()
    if disc_params:
        print(f"[*] Auto-discovered database config from: {config_path}")
        conn_params.update(disc_params)
        
    # Override with command line arguments if provided
    if args.host: conn_params['host'] = args.host
    if args.port: conn_params['port'] = args.port
    if args.dbname: conn_params['dbname'] = args.dbname
    if args.user: conn_params['user'] = args.user
    if args.password: conn_params['password'] = args.password
    
    # Default fallback if nothing discovered or provided
    if 'host' not in conn_params: conn_params['host'] = 'localhost'
    if 'port' not in conn_params: conn_params['port'] = 5432
    if 'dbname' not in conn_params: conn_params['dbname'] = 'muro_dev'
    if 'user' not in conn_params: conn_params['user'] = 'muro_user'
    if 'password' not in conn_params: conn_params['password'] = 'muro_pass_2024'

    print("Reading Excel file 'dereceuzem.xlsx'...")
    if not os.path.exists('dereceuzem.xlsx'):
        print("[!] Error: 'dereceuzem.xlsx' not found in current directory.")
        sys.exit(1)
        
    try:
        wb = openpyxl.load_workbook('dereceuzem.xlsx', data_only=True)
    except Exception as e:
        print(f"[!] Error opening Excel: {e}")
        sys.exit(1)
        
    if 'Sayfa1' not in wb.sheetnames or 'Sayfa2' not in wb.sheetnames:
        print("[!] Error: Excel must contain sheets: Sayfa1 and Sayfa2")
        sys.exit(1)
        
    s1 = wb['Sayfa1']
    s2 = wb['Sayfa2']
    
    rows1 = list(s1.iter_rows(values_only=True))
    rows2 = list(s2.iter_rows(values_only=True))
    
    # 1. Parse Courses
    courses_to_create = set()
    course_modes = {} 
    group_courses = [] 
    
    for r in rows2[1:]:
        g_name = r[3]
        c_name = r[4]
        mode = r[6]
        if g_name and c_name:
            g_name = str(g_name).strip()
            c_name = str(c_name).strip()
            mode = str(mode).strip()
            courses_to_create.add(c_name)
            course_modes[c_name] = mode
            group_courses.append((g_name, c_name, mode))
            
    # 2. Parse Groups
    groups_to_create = set()
    for r in rows2[1:]:
        g_name_left = r[3]
        g_name_right = r[10]
        if g_name_left:
            groups_to_create.add(str(g_name_left).strip())
        if g_name_right:
            groups_to_create.add(str(g_name_right).strip())
            
    # 3. Parse Students
    students_to_create = {} 
    
    for r in rows1[1:]:
        name = r[1]
        phone = r[2]
        email = r[3]
        if name or phone or email:
            cleaned_phone = clean_phone_number(phone)
            if cleaned_phone:
                students_to_create[cleaned_phone] = {
                    'name': str(name).strip(),
                    'phone': cleaned_phone,
                    'email': str(email).strip().lower() if email else f"student_{cleaned_phone}@dereceuzem.com"
                }
                
    student_groups = [] 
    for r in rows2[1:]:
        if len(r) > 14:
            g_name = r[10]
            name = r[12]
            phone = r[13]
            email = r[14]
            if name or phone or email:
                cleaned_phone = clean_phone_number(phone)
                if cleaned_phone:
                    g_name = str(g_name).strip()
                    student_groups.append((cleaned_phone, g_name))
                    if cleaned_phone not in students_to_create:
                        students_to_create[cleaned_phone] = {
                            'name': str(name).strip(),
                            'phone': cleaned_phone,
                            'email': str(email).strip().lower() if email else f"student_{cleaned_phone}@dereceuzem.com"
                        }
                        
    print(f"\n[+] Parsed Excel Data successfully:")
    print(f"    - Unique Groups  : {len(groups_to_create)}")
    print(f"    - Unique Courses : {len(courses_to_create)}")
    print(f"    - Unique Students: {len(students_to_create)}")
    print(f"    - Group-Course Mappings: {len(group_courses)}")
    print(f"    - Student-Group Mappings: {len(student_groups)}")
    
    if args.no_db:
        print("\n[*] Running in local parse-only mode. Skipping DB connection.")
        sys.exit(0)
        
    print(f"\n[*] Connecting to database '{conn_params.get('dbname')}' on {conn_params.get('host')}:{conn_params.get('port')}...")
    try:
        conn = psycopg2.connect(**conn_params)
        cur = conn.cursor()
    except Exception as e:
        print(f"[!] Database Connection Failed: {e}")
        print("[!] Please check if your SSH tunnel is active or DB credentials are correct.")
        sys.exit(1)
        
    print("[+] Database connected successfully!")
    
    # Retrieve existing records to verify duplicates
    print("[*] Retrieving existing database records...")
    
    cur.execute('SELECT "Id", "Title" FROM "Courses" WHERE "IsDeleted" = False;')
    existing_courses = {r[1].lower().strip(): r[0] for r in cur.fetchall()}
    
    cur.execute('SELECT "Id", "Name" FROM "Groups" WHERE "IsDeleted" = False;')
    existing_groups = {r[1].lower().strip(): r[0] for r in cur.fetchall()}
    
    cur.execute('SELECT "Id", "Email", "Phone", "Username" FROM "Users";')
    users_db = cur.fetchall()
    existing_users_by_email = {r[1].lower().strip(): r[0] for r in users_db if r[1]}
    existing_users_by_phone = {r[2].strip(): r[0] for r in users_db if r[2]}
    existing_users_by_username = {r[3].lower().strip(): r[0] for r in users_db if r[3]}
    
    cur.execute('SELECT "CourseId", "GroupId" FROM "CourseGroups";')
    existing_course_groups = set((r[0], r[1]) for r in cur.fetchall())
    
    cur.execute('SELECT "UserId", "GroupId" FROM "GroupMembers";')
    existing_group_members = set((r[0], r[1]) for r in cur.fetchall())
    
    print("[+] Existing records loaded.")
    
    # Calculate write plan
    planned_students_new = sum(1 for p in students_to_create if p not in existing_users_by_phone and p not in existing_users_by_username)
    planned_courses_new = sum(1 for c in courses_to_create if c.lower().strip() not in existing_courses)
    planned_groups_new = sum(1 for g in groups_to_create if g.lower().strip() not in existing_groups)
    
    print("\n" + "=" * 50)
    print("                  PROPOSED WORKPLAN")
    print("=" * 50)
    print(f" 1. Create Students   : {planned_students_new} new to create, {len(students_to_create) - planned_students_new} already exist")
    print(f" 2. Create Courses    : {planned_courses_new} new to create, {len(courses_to_create) - planned_courses_new} already exist")
    print(f" 3. Create Groups     : {planned_groups_new} new to create, {len(groups_to_create) - planned_groups_new} already exist")
    print(f" 4. Define Course-Group Mappings (CourseGroups)")
    print(f" 5. Define Student-Group Mappings (GroupMembers)")
    print("=" * 50)
    
    if not args.execute:
        print("\n[*] RUNNING IN DRY-RUN MODE. No changes will be written.")
        print("[*] To write changes to the database, run with: python import_data.py --execute")
        cur.close()
        conn.close()
        sys.exit(0)
        
    # Interactive confirmation prompt
    confirm = input("\n[?] Are you sure you want to execute this import on the database? (y/N): ")
    if confirm.lower().strip() not in ('y', 'yes'):
        print("[*] Operation cancelled.")
        cur.close()
        conn.close()
        sys.exit(0)
        
    print("\n[*] Starting transaction...")
    try:
        # STEP 1: CREATE STUDENTS (USERS)
        print("[1/5] Creating Students...")
        student_map = {}
        users_created = 0
        users_skipped = 0
        
        for phone, stud in students_to_create.items():
            email = stud['email']
            name = stud['name']
            first, last = parse_name(name)
            
            u_id = None
            if email.lower().strip() in existing_users_by_email:
                u_id = existing_users_by_email[email.lower().strip()]
            elif phone in existing_users_by_phone:
                u_id = existing_users_by_phone[phone]
            elif phone in existing_users_by_username:
                u_id = existing_users_by_username[phone]
                
            if u_id:
                student_map[phone] = u_id
                users_skipped += 1
            else:
                u_id = str(uuid.uuid4())
                student_map[phone] = u_id
                cur.execute(
                    'INSERT INTO "Users" ("Id", "FirstName", "LastName", "Email", "Username", "Phone", "PasswordHash", "Role", "StudentType", "IsActive", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
                    (u_id, first, last, email, phone, phone, email, 'Student', 'Active', True, False, datetime.datetime.utcnow())
                )
                users_created += 1
        print(f"      -> Created: {users_created}, Skipped: {users_skipped}")
        
        # STEP 2: CREATE COURSES
        print("[2/5] Creating Courses...")
        course_map = {}
        courses_created = 0
        courses_skipped = 0
        for c_name in courses_to_create:
            c_key = c_name.lower().strip()
            if c_key in existing_courses:
                course_map[c_name] = existing_courses[c_key]
                courses_skipped += 1
            else:
                course_id = str(uuid.uuid4())
                course_map[c_name] = course_id
                cur.execute(
                    'INSERT INTO "Courses" ("Id", "Title", "IsDeleted", "IsPublished", "CourseType", "Mode", "CreatedAt") VALUES (%s, %s, %s, %s, %s, %s, %s)',
                    (course_id, c_name, False, True, 'Online', 'Offline', datetime.datetime.utcnow())
                )
                courses_created += 1
        print(f"      -> Created: {courses_created}, Skipped: {courses_skipped}")
        
        # STEP 3: CREATE GROUPS
        print("[3/5] Creating Groups...")
        group_map = {}
        groups_created = 0
        groups_skipped = 0
        for g_name in groups_to_create:
            g_key = g_name.lower().strip()
            if g_key in existing_groups:
                group_map[g_name] = existing_groups[g_key]
                groups_skipped += 1
            else:
                group_id = str(uuid.uuid4())
                group_map[g_name] = group_id
                cur.execute(
                    'INSERT INTO "Groups" ("Id", "Name", "IsDeleted", "CreatedAt") VALUES (%s, %s, %s, %s)',
                    (group_id, g_name, False, datetime.datetime.utcnow())
                )
                groups_created += 1
        print(f"      -> Created: {groups_created}, Skipped: {groups_skipped}")
        
        # STEP 4: DEFINE COURSE-GROUP MAPPINGS (CourseGroups)
        print("[4/5] Defining Course-Group Mappings...")
        cg_created = 0
        cg_skipped = 0
        for g_name, c_name, mode in group_courses:
            g_id = group_map.get(g_name)
            c_id = course_map.get(c_name)
            if g_id and c_id:
                if (c_id, g_id) in existing_course_groups:
                    cg_skipped += 1
                else:
                    c_mode = 'Online' if mode == 'Canlı' else 'Offline'
                    cur.execute(
                        'INSERT INTO "CourseGroups" ("Id", "CourseId", "GroupId", "Mode", "AssignedAt") VALUES (%s, %s, %s, %s, %s)',
                        (str(uuid.uuid4()), c_id, g_id, c_mode, datetime.datetime.utcnow())
                    )
                    cg_created += 1
        print(f"      -> Created: {cg_created}, Skipped: {cg_skipped}")
        
        # STEP 5: DEFINE STUDENT-GROUP MAPPINGS (GroupMembers)
        print("[5/5] Defining Student-Group Mappings...")
        members_created = 0
        members_skipped = 0
        for phone, g_name in student_groups:
            u_id = student_map.get(phone)
            g_id = group_map.get(g_name)
            if u_id and g_id:
                if (u_id, g_id) in existing_group_members:
                    members_skipped += 1
                else:
                    cur.execute(
                        'INSERT INTO "GroupMembers" ("Id", "UserId", "GroupId", "Role", "Status", "AddedAt") VALUES (%s, %s, %s, %s, %s, %s)',
                        (str(uuid.uuid4()), u_id, g_id, 0, 'active', datetime.datetime.utcnow())
                    )
                    members_created += 1
        print(f"      -> Created: {members_created}, Skipped: {members_skipped}")
        
        # Commit transaction
        print("\n[*] Committing database transaction...")
        conn.commit()
        print("[+] SUCCESS: Data import completed successfully and committed to the database!")
        
    except Exception as db_err:
        print(f"\n[!] Database Transaction Error encountered: {db_err}")
        print("[!] Rolling back all changes in this transaction to preserve data integrity...")
        conn.rollback()
        print("[*] Rollback complete. No changes were saved to the database.")
        sys.exit(1)
        
    finally:
        cur.close()
        conn.close()

if __name__ == '__main__':
    main()
